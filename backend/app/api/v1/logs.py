from typing import Optional
from fastapi import APIRouter, Depends, Query, Request, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, text
from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from app.core.database import get_db, SessionLocal
from app.models import OperationLog, LoginLog, AccessLog, UserProfile
from app.utils.permissions import require_permission
from app.utils.timezone import get_db_now, get_today_start, to_local
from app.services.log_service import LogService
from app.utils.cache import cache_manager
from pydantic import BaseModel, field_validator
import uuid
from typing import Dict

router = APIRouter(prefix="/logs", tags=["Logs"])

CACHE_NAME = "logs"
CACHE_TTL_STATS = 60

export_progress: Dict[str, Dict] = {}
active_exports = 0
MAX_CONCURRENT_EXPORTS = 3


class OperationLogItem(BaseModel):
    id: int
    username: Optional[str]
    user_id: Optional[int]
    avatar_type: Optional[str] = None
    avatar_url: Optional[str] = None
    avatar_gradient: Optional[list] = None
    oauth_avatar_url: Optional[str] = None
    action: str
    module: str
    description: Optional[str]
    target_type: Optional[str]
    target_id: Optional[int]
    request_method: Optional[str]
    request_url: Optional[str]
    ip_address: Optional[str]
    status: str
    created_at: Optional[str] = None
    
    @field_validator('created_at', mode='before')
    @classmethod
    def serialize_created_at(cls, v):
        if v is None:
            return None
        local_dt = to_local(v)
        return local_dt.isoformat()
    
    class Config:
        from_attributes = True


class LoginLogItem(BaseModel):
    id: int
    username: Optional[str]
    user_id: Optional[int]
    avatar_type: Optional[str] = None
    avatar_url: Optional[str] = None
    avatar_gradient: Optional[list] = None
    oauth_avatar_url: Optional[str] = None
    login_type: str
    ip_address: Optional[str]
    location: Optional[str]
    browser: Optional[str]
    os: Optional[str]
    device: Optional[str]
    status: str
    fail_reason: Optional[str]
    created_at: Optional[str] = None
    
    @field_validator('created_at', mode='before')
    @classmethod
    def serialize_created_at(cls, v):
        if v is None:
            return None
        local_dt = to_local(v)
        return local_dt.isoformat()
    
    class Config:
        from_attributes = True


class AccessLogItem(BaseModel):
    id: int
    username: Optional[str]
    user_id: Optional[int]
    avatar_type: Optional[str] = None
    avatar_url: Optional[str] = None
    avatar_gradient: Optional[list] = None
    oauth_avatar_url: Optional[str] = None
    request_method: Optional[str]
    request_path: Optional[str]
    response_status: Optional[int]
    response_time: Optional[float]
    ip_address: Optional[str]
    created_at: Optional[str] = None
    
    @field_validator('created_at', mode='before')
    @classmethod
    def serialize_created_at(cls, v):
        if v is None:
            return None
        local_dt = to_local(v)
        return local_dt.isoformat()
    
    class Config:
        from_attributes = True


class LogStats(BaseModel):
    total_operations: int
    total_logins: int
    total_access: int
    today_operations: int
    today_logins: int
    failed_logins: int


@router.get("/stats", response_model=LogStats)
async def get_log_stats(
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    cached = cache_manager.get(CACHE_NAME, "stats")
    if cached:
        return LogStats(**cached)
    
    today_start = get_today_start()
    
    total_operations = db.query(func.count(OperationLog.id)).scalar() or 0
    total_logins = db.query(func.count(LoginLog.id)).scalar() or 0
    total_access = db.query(func.count(AccessLog.id)).scalar() or 0
    today_operations = db.query(func.count(OperationLog.id)).filter(OperationLog.created_at >= today_start).scalar() or 0
    today_logins = db.query(func.count(LoginLog.id)).filter(LoginLog.created_at >= today_start).scalar() or 0
    failed_logins = db.query(func.count(LoginLog.id)).filter(LoginLog.status == 'failed').scalar() or 0
    
    result = {
        "total_operations": total_operations,
        "total_logins": total_logins,
        "total_access": total_access,
        "today_operations": today_operations,
        "today_logins": today_logins,
        "failed_logins": failed_logins,
    }
    
    cache_manager.set(CACHE_NAME, "stats", result, ttl=CACHE_TTL_STATS)
    return LogStats(**result)


@router.get("/operations", response_model=dict)
async def get_operation_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    module: Optional[str] = None,
    action: Optional[str] = None,
    description: Optional[str] = None,
    status: Optional[str] = None,
    username: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    query = db.query(OperationLog)
    
    if module:
        query = query.filter(OperationLog.module.contains(module))
    if action:
        query = query.filter(OperationLog.action.contains(action))
    if description:
        query = query.filter(OperationLog.description.ilike(f"%{description}%"))
    if status:
        query = query.filter(OperationLog.status == status)
    if username:
        query = query.filter(OperationLog.username.contains(username))
    if start_date:
        query = query.filter(OperationLog.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        end_datetime = datetime.fromisoformat(end_date)
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(OperationLog.created_at <= end_datetime)
    
    total = query.count()
    logs = query.order_by(desc(OperationLog.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    
    user_ids = set()
    for log in logs:
        if log.user_id:
            user_ids.add(log.user_id)
    
    user_profiles = {}
    if user_ids:
        profiles = db.query(UserProfile).filter(UserProfile.user_id.in_(user_ids)).all()
        user_profiles = {p.user_id: p for p in profiles}
    
    items = []
    for log in logs:
        avatar_type = None
        avatar_url = None
        avatar_gradient = None
        oauth_avatar_url = None
        
        if log.user_id:
            profile = user_profiles.get(log.user_id)
            if profile:
                avatar_type = profile.avatar_type.value if profile.avatar_type else None
                avatar_url = profile.avatar_url
                avatar_gradient = profile.default_avatar_gradient
                oauth_avatar_url = profile.oauth_avatar_url
        
        items.append(OperationLogItem(
            id=log.id,
            username=log.username,
            user_id=log.user_id,
            avatar_type=avatar_type,
            avatar_url=avatar_url,
            avatar_gradient=avatar_gradient,
            oauth_avatar_url=oauth_avatar_url,
            action=log.action,
            module=log.module,
            description=log.description,
            target_type=log.target_type,
            target_id=log.target_id,
            request_method=log.request_method,
            request_url=log.request_url,
            ip_address=log.ip_address,
            status=log.status,
            created_at=log.created_at
        ))
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.get("/logins", response_model=dict)
async def get_login_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    login_type: Optional[str] = None,
    status: Optional[str] = None,
    username: Optional[str] = None,
    ip_address: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    query = db.query(LoginLog)
    
    if login_type:
        query = query.filter(LoginLog.login_type.contains(login_type))
    if status:
        query = query.filter(LoginLog.status == status)
    if username:
        query = query.filter(LoginLog.username.contains(username))
    if ip_address:
        query = query.filter(LoginLog.ip_address.contains(ip_address))
    if start_date:
        query = query.filter(LoginLog.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        end_datetime = datetime.fromisoformat(end_date)
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(LoginLog.created_at <= end_datetime)
    
    total = query.count()
    logs = query.order_by(desc(LoginLog.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    
    user_ids = set()
    for log in logs:
        if log.user_id:
            user_ids.add(log.user_id)
    
    user_profiles = {}
    if user_ids:
        profiles = db.query(UserProfile).filter(UserProfile.user_id.in_(user_ids)).all()
        user_profiles = {p.user_id: p for p in profiles}
    
    items = []
    for log in logs:
        avatar_type = None
        avatar_url = None
        avatar_gradient = None
        oauth_avatar_url = None
        
        if log.user_id:
            profile = user_profiles.get(log.user_id)
            if profile:
                avatar_type = profile.avatar_type.value if profile.avatar_type else None
                avatar_url = profile.avatar_url
                avatar_gradient = profile.default_avatar_gradient
                oauth_avatar_url = profile.oauth_avatar_url
        
        items.append(LoginLogItem(
            id=log.id,
            username=log.username,
            user_id=log.user_id,
            avatar_type=avatar_type,
            avatar_url=avatar_url,
            avatar_gradient=avatar_gradient,
            oauth_avatar_url=oauth_avatar_url,
            login_type=log.login_type,
            ip_address=log.ip_address,
            location=log.location,
            browser=log.browser,
            os=log.os,
            device=log.device,
            status=log.status,
            fail_reason=log.fail_reason,
            created_at=log.created_at
        ))
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.get("/access", response_model=dict)
async def get_access_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    username: Optional[str] = None,
    request_method: Optional[str] = None,
    path: Optional[str] = None,
    ip_address: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    query = db.query(AccessLog)
    
    if username:
        if username == "__guest__":
            query = query.filter(AccessLog.username.is_(None) | (AccessLog.username == ""))
        else:
            query = query.filter(AccessLog.username.contains(username))
    if request_method:
        query = query.filter(AccessLog.request_method.contains(request_method))
    if path:
        query = query.filter(AccessLog.request_path.contains(path))
    if ip_address:
        query = query.filter(AccessLog.ip_address.contains(ip_address))
    if start_date:
        query = query.filter(AccessLog.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        end_datetime = datetime.fromisoformat(end_date)
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(AccessLog.created_at <= end_datetime)
    
    total = query.count()
    logs = query.order_by(desc(AccessLog.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    
    user_ids = set()
    for log in logs:
        if log.user_id:
            user_ids.add(log.user_id)
    
    user_profiles = {}
    if user_ids:
        profiles = db.query(UserProfile).filter(UserProfile.user_id.in_(user_ids)).all()
        user_profiles = {p.user_id: p for p in profiles}
    
    items = []
    for log in logs:
        avatar_type = None
        avatar_url = None
        avatar_gradient = None
        oauth_avatar_url = None
        
        if log.user_id:
            profile = user_profiles.get(log.user_id)
            if profile:
                avatar_type = profile.avatar_type.value if profile.avatar_type else None
                avatar_url = profile.avatar_url
                avatar_gradient = profile.default_avatar_gradient
                oauth_avatar_url = profile.oauth_avatar_url
        
        items.append(AccessLogItem(
            id=log.id,
            username=log.username,
            user_id=log.user_id,
            avatar_type=avatar_type,
            avatar_url=avatar_url,
            avatar_gradient=avatar_gradient,
            oauth_avatar_url=oauth_avatar_url,
            request_method=log.request_method,
            request_path=log.request_path,
            response_status=log.response_status,
            response_time=log.response_time,
            ip_address=log.ip_address,
            created_at=log.created_at
        ))
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.delete("/operations/clear")
async def clear_operation_logs(
    days: int = Query(30, ge=1),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("log.clear"))
):
    cutoff_date = get_db_now() - timedelta(days=days)
    deleted = db.query(OperationLog).filter(OperationLog.created_at < cutoff_date).delete()
    db.commit()
    
    LogService.log_operation(
        db=db,
        user_id=current_user.id if hasattr(current_user, 'id') else None,
        username=current_user.username if hasattr(current_user, 'username') else None,
        action="clear",
        module="operation_log",
        description=f"清理操作日志: 删除{deleted}条 ({days}天前)",
        request=request
    )
    
    return {"message": f"已删除 {deleted} 条操作日志"}


@router.delete("/logins/clear")
async def clear_login_logs(
    days: int = Query(30, ge=1),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("log.clear"))
):
    cutoff_date = get_db_now() - timedelta(days=days)
    deleted = db.query(LoginLog).filter(LoginLog.created_at < cutoff_date).delete()
    db.commit()
    
    LogService.log_operation(
        db=db,
        user_id=current_user.id if hasattr(current_user, 'id') else None,
        username=current_user.username if hasattr(current_user, 'username') else None,
        action="clear",
        module="login_log",
        description=f"清理登录日志: 删除{deleted}条 ({days}天前)",
        request=request
    )
    
    return {"message": f"已删除 {deleted} 条登录日志"}


@router.delete("/access/clear")
async def clear_access_logs(
    days: int = Query(30, ge=1),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("log.clear"))
):
    cutoff_date = get_db_now() - timedelta(days=days)
    deleted = db.query(AccessLog).filter(AccessLog.created_at < cutoff_date).delete()
    db.commit()
    
    LogService.log_operation(
        db=db,
        user_id=current_user.id if hasattr(current_user, 'id') else None,
        username=current_user.username if hasattr(current_user, 'username') else None,
        action="clear",
        module="access_log",
        description=f"清理访问日志: 删除{deleted}条 ({days}天前)",
        request=request
    )
    
    return {"message": f"已删除 {deleted} 条访问日志"}


def create_excel_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "系统日志"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    return wb, ws, header_font, header_fill, header_alignment, cell_alignment, thin_border


def apply_header_style(ws, headers, row_num, header_font, header_fill, header_alignment, thin_border):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def apply_cell_style(ws, row_data, row_num, cell_alignment, thin_border):
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border


def auto_adjust_column_width(ws, column_count):
    for col in range(1, column_count + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = min(cell_length, 50)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = max(adjusted_width, 10)


@router.get("/export/progress/{task_id}")
async def get_export_progress(
    task_id: str
):
    progress = export_progress.get(task_id, {"current": 0, "total": 0, "cancelled": False})
    return progress


@router.post("/export/cancel/{task_id}")
async def cancel_export_task(
    task_id: str
):
    if task_id in export_progress:
        export_progress[task_id]["cancelled"] = True
    return {"message": "Export cancelled"}


@router.delete("/export/progress/{task_id}")
async def clear_export_progress(
    task_id: str
):
    if task_id in export_progress:
        del export_progress[task_id]
    return {"message": "Progress cleared"}


@router.get("/export/operations/count")
async def get_operation_logs_export_count(
    task_id: Optional[str] = None,
    module: Optional[str] = None,
    action: Optional[str] = None,
    status: Optional[str] = None,
    username: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    query = db.query(OperationLog)
    
    if module:
        query = query.filter(OperationLog.module.contains(module))
    if action:
        query = query.filter(OperationLog.action.contains(action))
    if status:
        query = query.filter(OperationLog.status == status)
    if username:
        query = query.filter(OperationLog.username.contains(username))
    if start_date:
        query = query.filter(OperationLog.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        end_datetime = datetime.fromisoformat(end_date)
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(OperationLog.created_at <= end_datetime)
    
    total_count = query.count()
    
    if task_id:
        export_progress[task_id] = {"current": 0, "total": total_count, "cancelled": False, "stage": "pending"}
    
    return {"total": total_count}


@router.get("/export/operations")
async def export_operation_logs(
    task_id: Optional[str] = None,
    module: Optional[str] = None,
    action: Optional[str] = None,
    status: Optional[str] = None,
    username: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    global active_exports
    
    if active_exports >= MAX_CONCURRENT_EXPORTS:
        raise HTTPException(status_code=429, detail="Too many concurrent exports. Please try again later.")
    
    active_exports += 1
    
    query = db.query(OperationLog)
    
    if module:
        query = query.filter(OperationLog.module.contains(module))
    if action:
        query = query.filter(OperationLog.action.contains(action))
    if status:
        query = query.filter(OperationLog.status == status)
    if username:
        query = query.filter(OperationLog.username.contains(username))
    if start_date:
        query = query.filter(OperationLog.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        end_datetime = datetime.fromisoformat(end_date)
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(OperationLog.created_at <= end_datetime)
    
    if task_id and task_id in export_progress:
        total_count = export_progress[task_id]["total"]
    else:
        total_count = query.count()
        if task_id:
            export_progress[task_id] = {"current": 0, "total": total_count, "cancelled": False, "stage": "processing"}
    
    batch_size = 5000
    
    if task_id and task_id in export_progress:
        export_progress[task_id]["stage"] = "processing"
    
    def generate_excel():
        global active_exports
        try:
            wb, ws, header_font, header_fill, header_alignment, cell_alignment, thin_border = create_excel_workbook()
            ws.title = "操作日志"
            
            headers = ["日志ID", "发生时间", "操作人", "模块", "操作", "描述", "IP地址", "状态"]
            apply_header_style(ws, headers, 1, header_font, header_fill, header_alignment, thin_border)
            
            current_row = 2
            offset = 0
            processed_count = 0
            
            while offset < total_count:
                if task_id and task_id in export_progress and export_progress[task_id].get("cancelled", False):
                    return
                
                session = SessionLocal()
                try:
                    batch_query = session.query(OperationLog)
                    
                    if module:
                        batch_query = batch_query.filter(OperationLog.module.contains(module))
                    if action:
                        batch_query = batch_query.filter(OperationLog.action.contains(action))
                    if status:
                        batch_query = batch_query.filter(OperationLog.status == status)
                    if username:
                        batch_query = batch_query.filter(OperationLog.username.contains(username))
                    if start_date:
                        batch_query = batch_query.filter(OperationLog.created_at >= datetime.fromisoformat(start_date))
                    if end_date:
                        end_datetime = datetime.fromisoformat(end_date)
                        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
                        batch_query = batch_query.filter(OperationLog.created_at <= end_datetime)
                    
                    batch_logs = batch_query.order_by(desc(OperationLog.created_at)).offset(offset).limit(batch_size).all()
                    
                    for log in batch_logs:
                        created_at_str = to_local(log.created_at).strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
                        status_str = "成功" if log.status == "success" else "失败"
                        row_data = [
                            log.id,
                            created_at_str,
                            log.username or "-",
                            log.module,
                            log.action,
                            log.description or "-",
                            log.ip_address or "-",
                            status_str
                        ]
                        apply_cell_style(ws, row_data, current_row, cell_alignment, thin_border)
                        current_row += 1
                        processed_count += 1
                finally:
                    session.close()
                
                if task_id and task_id in export_progress:
                    export_progress[task_id]["current"] = processed_count
                
                offset += batch_size
            
            auto_adjust_column_width(ws, len(headers))
            
            if task_id and task_id in export_progress:
                export_progress[task_id]["stage"] = "generating"
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            if task_id and task_id in export_progress:
                export_progress[task_id]["stage"] = "streaming"
            
            chunk_size = 65536
            while True:
                chunk = output.read(chunk_size)
                if not chunk:
                    break
                yield chunk
        finally:
            active_exports -= 1
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"操作日志_{timestamp}.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        generate_excel(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "X-Total-Count": str(total_count)
        }
    )


@router.get("/export/logins/count")
async def get_login_logs_export_count(
    task_id: Optional[str] = None,
    login_type: Optional[str] = None,
    status: Optional[str] = None,
    username: Optional[str] = None,
    ip_address: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    query = db.query(LoginLog)
    
    if login_type:
        query = query.filter(LoginLog.login_type.contains(login_type))
    if status:
        query = query.filter(LoginLog.status == status)
    if username:
        query = query.filter(LoginLog.username.contains(username))
    if ip_address:
        query = query.filter(LoginLog.ip_address.contains(ip_address))
    if start_date:
        query = query.filter(LoginLog.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        end_datetime = datetime.fromisoformat(end_date)
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(LoginLog.created_at <= end_datetime)
    
    total_count = query.count()
    
    if task_id:
        export_progress[task_id] = {"current": 0, "total": total_count, "cancelled": False, "stage": "pending"}
    
    return {"total": total_count}


@router.get("/export/logins")
async def export_login_logs(
    task_id: Optional[str] = None,
    login_type: Optional[str] = None,
    status: Optional[str] = None,
    username: Optional[str] = None,
    ip_address: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    query = db.query(LoginLog)
    
    if login_type:
        query = query.filter(LoginLog.login_type.contains(login_type))
    if status:
        query = query.filter(LoginLog.status == status)
    if username:
        query = query.filter(LoginLog.username.contains(username))
    if ip_address:
        query = query.filter(LoginLog.ip_address.contains(ip_address))
    if start_date:
        query = query.filter(LoginLog.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        end_datetime = datetime.fromisoformat(end_date)
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(LoginLog.created_at <= end_datetime)
    
    if task_id and task_id in export_progress:
        total_count = export_progress[task_id]["total"]
    else:
        total_count = query.count()
        if task_id:
            export_progress[task_id] = {"current": 0, "total": total_count, "cancelled": False, "stage": "processing"}
    
    batch_size = 5000
    
    if task_id and task_id in export_progress:
        export_progress[task_id]["stage"] = "processing"
    
    def generate_excel():
        wb, ws, header_font, header_fill, header_alignment, cell_alignment, thin_border = create_excel_workbook()
        ws.title = "登录日志"
        
        headers = ["日志ID", "发生时间", "用户", "登录类型", "IP地址", "浏览器", "操作系统", "状态", "失败原因"]
        apply_header_style(ws, headers, 1, header_font, header_fill, header_alignment, thin_border)
        
        current_row = 2
        offset = 0
        processed_count = 0
        
        while offset < total_count:
            if task_id and task_id in export_progress and export_progress[task_id].get("cancelled", False):
                return
            
            session = SessionLocal()
            try:
                batch_query = session.query(LoginLog)
                
                if login_type:
                    batch_query = batch_query.filter(LoginLog.login_type.contains(login_type))
                if status:
                    batch_query = batch_query.filter(LoginLog.status == status)
                if username:
                    batch_query = batch_query.filter(LoginLog.username.contains(username))
                if ip_address:
                    batch_query = batch_query.filter(LoginLog.ip_address.contains(ip_address))
                if start_date:
                    batch_query = batch_query.filter(LoginLog.created_at >= datetime.fromisoformat(start_date))
                if end_date:
                    end_datetime = datetime.fromisoformat(end_date)
                    end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
                    batch_query = batch_query.filter(LoginLog.created_at <= end_datetime)
                
                batch_logs = batch_query.order_by(desc(LoginLog.created_at)).offset(offset).limit(batch_size).all()
                
                for log in batch_logs:
                    created_at_str = to_local(log.created_at).strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
                    status_str = "成功" if log.status == "success" else "失败"
                    row_data = [
                        log.id,
                        created_at_str,
                        log.username or "-",
                        log.login_type,
                        log.ip_address or "-",
                        log.browser or "-",
                        log.os or "-",
                        status_str,
                        log.fail_reason or "-"
                    ]
                    apply_cell_style(ws, row_data, current_row, cell_alignment, thin_border)
                    current_row += 1
                    processed_count += 1
            finally:
                session.close()
            
            if task_id and task_id in export_progress:
                export_progress[task_id]["current"] = processed_count
            
            offset += batch_size
        
        auto_adjust_column_width(ws, len(headers))
        
        if task_id and task_id in export_progress:
            export_progress[task_id]["stage"] = "generating"
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        if task_id and task_id in export_progress:
            export_progress[task_id]["stage"] = "streaming"
        
        chunk_size = 65536
        while True:
            chunk = output.read(chunk_size)
            if not chunk:
                break
            yield chunk
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"登录日志_{timestamp}.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        generate_excel(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "X-Total-Count": str(total_count)
        }
    )


@router.get("/export/access/count")
async def get_access_logs_export_count(
    task_id: Optional[str] = None,
    username: Optional[str] = None,
    request_method: Optional[str] = None,
    path: Optional[str] = None,
    ip_address: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    query = db.query(AccessLog)
    
    if username:
        if username == "__guest__":
            query = query.filter(AccessLog.username.is_(None) | (AccessLog.username == ""))
        else:
            query = query.filter(AccessLog.username.contains(username))
    if request_method:
        query = query.filter(AccessLog.request_method.contains(request_method))
    if path:
        query = query.filter(AccessLog.request_path.contains(path))
    if ip_address:
        query = query.filter(AccessLog.ip_address.contains(ip_address))
    if start_date:
        query = query.filter(AccessLog.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        end_datetime = datetime.fromisoformat(end_date)
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(AccessLog.created_at <= end_datetime)
    
    total_count = query.count()
    
    if task_id:
        export_progress[task_id] = {"current": 0, "total": total_count, "cancelled": False, "stage": "pending"}
    
    return {"total": total_count}


@router.get("/export/access")
async def export_access_logs(
    task_id: Optional[str] = None,
    username: Optional[str] = None,
    request_method: Optional[str] = None,
    path: Optional[str] = None,
    ip_address: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_permission("log.view"))
):
    query = db.query(AccessLog)
    
    if username:
        if username == "__guest__":
            query = query.filter(AccessLog.username.is_(None) | (AccessLog.username == ""))
        else:
            query = query.filter(AccessLog.username.contains(username))
    if request_method:
        query = query.filter(AccessLog.request_method.contains(request_method))
    if path:
        query = query.filter(AccessLog.request_path.contains(path))
    if ip_address:
        query = query.filter(AccessLog.ip_address.contains(ip_address))
    if start_date:
        query = query.filter(AccessLog.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        end_datetime = datetime.fromisoformat(end_date)
        end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
        query = query.filter(AccessLog.created_at <= end_datetime)
    
    if task_id and task_id in export_progress:
        total_count = export_progress[task_id]["total"]
    else:
        total_count = query.count()
        if task_id:
            export_progress[task_id] = {"current": 0, "total": total_count, "cancelled": False, "stage": "processing"}
    
    batch_size = 5000
    
    if task_id and task_id in export_progress:
        export_progress[task_id]["stage"] = "processing"
    
    def generate_excel():
        wb, ws, header_font, header_fill, header_alignment, cell_alignment, thin_border = create_excel_workbook()
        ws.title = "访问日志"
        
        headers = ["日志ID", "发生时间", "用户", "请求方法", "请求路径", "状态码", "响应时间(ms)", "IP地址"]
        apply_header_style(ws, headers, 1, header_font, header_fill, header_alignment, thin_border)
        
        current_row = 2
        offset = 0
        processed_count = 0
        
        while offset < total_count:
            if task_id and task_id in export_progress and export_progress[task_id].get("cancelled", False):
                return
            
            session = SessionLocal()
            try:
                batch_query = session.query(AccessLog)
                
                if username:
                    if username == "__guest__":
                        batch_query = batch_query.filter(AccessLog.username.is_(None) | (AccessLog.username == ""))
                    else:
                        batch_query = batch_query.filter(AccessLog.username.contains(username))
                if request_method:
                    batch_query = batch_query.filter(AccessLog.request_method.contains(request_method))
                if path:
                    batch_query = batch_query.filter(AccessLog.request_path.contains(path))
                if ip_address:
                    batch_query = batch_query.filter(AccessLog.ip_address.contains(ip_address))
                if start_date:
                    batch_query = batch_query.filter(AccessLog.created_at >= datetime.fromisoformat(start_date))
                if end_date:
                    end_datetime = datetime.fromisoformat(end_date)
                    end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
                    batch_query = batch_query.filter(AccessLog.created_at <= end_datetime)
                
                batch_logs = batch_query.order_by(desc(AccessLog.created_at)).offset(offset).limit(batch_size).all()
                
                for log in batch_logs:
                    created_at_str = to_local(log.created_at).strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
                    row_data = [
                        log.id,
                        created_at_str,
                        log.username or "游客",
                        log.request_method or "-",
                        log.request_path or "-",
                        log.response_status or "-",
                        f"{log.response_time:.2f}" if log.response_time else "-",
                        log.ip_address or "-"
                    ]
                    apply_cell_style(ws, row_data, current_row, cell_alignment, thin_border)
                    current_row += 1
                    processed_count += 1
            finally:
                session.close()
            
            if task_id and task_id in export_progress:
                export_progress[task_id]["current"] = processed_count
            
            offset += batch_size
        
        auto_adjust_column_width(ws, len(headers))
        
        if task_id and task_id in export_progress:
            export_progress[task_id]["stage"] = "generating"
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        if task_id and task_id in export_progress:
            export_progress[task_id]["stage"] = "streaming"
        
        chunk_size = 65536
        while True:
            chunk = output.read(chunk_size)
            if not chunk:
                break
            yield chunk
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"访问日志_{timestamp}.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        generate_excel(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "X-Total-Count": str(total_count)
        }
    )
